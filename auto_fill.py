import os
import sqlite3
import pandas as pd
import warnings
from openpyxl import load_workbook

# Ẩn các cảnh báo của openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

DB_NAME = "master_data.db"
TABLE_NAME = "master_data"


def show_menu():
    print("\n=== MENU CHÍNH ===")
    print("1. Khởi tạo master data")
    print("2. Fill data")
    print("0. Thoát")
    return input("Nhập lựa chọn của bạn: ").strip()


def init_master_data():
    try:
        file_path = input("Nhập hoặc kéo thả file master data Excel vào đây: ").strip().strip('"').strip("'")

        if not os.path.exists(file_path):
            print("❌ File không tồn tại.")
            return

        print("Đang đọc file Excel...")
        excel = pd.ExcelFile(file_path)
        all_data = []

        for sheet_name in excel.sheet_names:
            df = excel.parse(sheet_name)

            required_cols = {
                'Carrier Bkg no.': 'D',
                'Job no.': 'E',
                'Shipper Name': 'F',
                'Carrier': 'N',
                'Vessel name': 'AD',
                'Origin ETD': 'AF',
                'New ETD': 'AG',
                'Status': 'H'
            }

            missing = [c for c in required_cols.keys() if c not in df.columns]
            if missing:
                print(f"⚠️ Sheet '{sheet_name}' bị thiếu cột: {missing}, bỏ qua.")
                continue

            # Ép kiểu & lọc dữ liệu
            df['Status'] = df['Status'].astype(str).str.lower()
            df = df[~df['Status'].isin(['cancel', 'shipper cancel'])]
            df = df[df['Carrier Bkg no.'].notna() & (df['Carrier Bkg no.'].astype(str).str.strip() != '')]

            if not df.empty:
                selected = df[list(required_cols.keys())]
                all_data.append(selected)

        if not all_data:
            print("❌ Không có dữ liệu hợp lệ trong file.")
            return

        master_df = pd.concat(all_data, ignore_index=True)
        master_df = master_df.fillna('').astype(str)

        # Ghi dữ liệu vào SQLite
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute(f"DROP TABLE IF EXISTS {TABLE_NAME}")
        conn.commit()

        master_df.to_sql(TABLE_NAME, conn, index=False, if_exists='replace', dtype={col: 'TEXT' for col in master_df.columns})
        cursor.execute(f"CREATE INDEX IF NOT EXISTS idx_carrier_bkg_no ON {TABLE_NAME}(`Carrier Bkg no.`)")
        conn.commit()
        conn.close()

        print("✅ Khởi tạo master data hoàn tất.")

    except KeyboardInterrupt:
        print("\n⚠️ Đã hủy quá trình khởi tạo (Ctrl + C). Quay lại menu...")
    except Exception as e:
        print(f"❌ Lỗi trong quá trình init master data: {e}")

import os
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime


def normalize_date(value):
    """Chuẩn hóa giá trị ngày về dạng dd/mm/yy nếu có thể."""
    if value in (None, "", "NaT"):
        return ""
    
    # Nếu là kiểu datetime hoặc pandas Timestamp
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.strftime("%d/%m/%y")
    
    value = str(value).strip()
    if not value:
        return ""

    # Loại bỏ phần giờ nếu có
    if " " in value:
        value = value.split(" ")[0]

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            dt = datetime.strptime(value, fmt)
            return dt.strftime("%d/%m/%y")
        except ValueError:
            continue
    return value



def fill_data():
    DB_NAME = "master_data.db"
    TABLE_NAME = "master_data"

    try:
        # Kiểm tra master data
        if not os.path.exists(DB_NAME):
            print("⚠️ Chưa có master data. Hãy khởi tạo trước (chọn option 1).")
            return

        file_path = input("Nhập hoặc kéo thả file data cần xử lý: ").strip().strip('"').strip("'")
        if not os.path.exists(file_path):
            print("❌ File không tồn tại.")
            return

        # Đọc tất cả sheet
        excel = pd.ExcelFile(file_path)
        print("\n📘 Danh sách sheet có trong file:")
        for i, sheet in enumerate(excel.sheet_names, start=1):
            print(f"{i}. {sheet}")

        choice = input("\nNhập số sheet bạn muốn xử lý: ").strip()
        if not choice.isdigit() or int(choice) < 1 or int(choice) > len(excel.sheet_names):
            print("❌ Lựa chọn không hợp lệ.")
            return

        sheet_name = excel.sheet_names[int(choice) - 1]
        print(f"👉 Đang xử lý sheet: {sheet_name}")

        # Đọc sheet được chọn
        input_df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str).fillna('')
        if 'BK #' not in input_df.columns:
            print("❌ File không có cột 'BK #'.")
            return

        # Kết nối DB
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()

        # Load workbook (để ghi đè mà giữ nguyên format)
        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        # Lấy index của cột cần fill
        col_map = {
            "LINE": "Carrier",
            "JOB": "Job no.",
            "VESSEL": "Vessel name",
            "BRANCH": "Shipper Name",
            "ETD": None  # xử lý riêng
        }

        header_row = 1
        col_index = {cell.value.strip(): cell.column for cell in ws[header_row] if cell.value}

        # Duyệt từng dòng
        for i, bk_value in enumerate(input_df['BK #'], start=2):  # Dòng bắt đầu từ 2 (sau header)
            bk_value = str(bk_value).strip()
            if not bk_value:
                continue  # bỏ qua BK# rỗng

            # Truy vấn trong master data
            query = f"SELECT * FROM {TABLE_NAME} WHERE `Carrier Bkg no.` = ?"
            rows = cursor.execute(query, (bk_value,)).fetchall()

            if len(rows) == 0:
                print(f"⚠️ BK # : {bk_value} không tìm thấy trong master data.")
                continue
            elif len(rows) > 1:
                print(f"⚠️ BK # : {bk_value} đang bị trùng lặp ở master data, phát hiện có {len(rows)} rows bị trùng. Đang sử dụng row đầu tiên để fill data.")

            # Lấy dòng đầu tiên
            columns = [desc[0] for desc in cursor.description]
            data = dict(zip(columns, rows[0]))

            # Xử lý ETD
            origin_etd = str(data.get("Origin ETD", "")).strip()
            new_etd = str(data.get("New ETD", "")).strip()

            if origin_etd and not new_etd:
                etd_value = normalize_date(origin_etd)
            elif origin_etd and new_etd:
                etd_value = normalize_date(new_etd)
            else:
                etd_value = ""

            # Fill các cột tương ứng
            for col_name, db_field in col_map.items():
                if col_name not in col_index:
                    continue  # bỏ qua nếu cột không tồn tại trong input
                col_num = col_index[col_name]
                cell = ws.cell(row=i, column=col_num)
                if col_name == "ETD":
                    cell.value = etd_value
                else:
                    cell.value = data.get(db_field, "")

        # Lưu lại file gốc (giữ nguyên format)
        wb.save(file_path)
        wb.close()
        conn.close()

        print("\n✅ Fill data hoàn tất.")

    except KeyboardInterrupt:
        print("\n🛑 Đã dừng chương trình theo yêu cầu.")
    except Exception as e:
        print(f"❌ Lỗi trong quá trình fill data: {e}")


def main():
    try:
        while True:
            choice = show_menu()
            if choice == "1":
                init_master_data()
            elif choice == "2":
                fill_data()
            elif choice == "0":
                print("👋 Tạm biệt!")
                break
            else:
                print("❌ Lựa chọn không hợp lệ, vui lòng nhập lại.")
    except KeyboardInterrupt:
        print("\n👋 Tạm biệt!")


if __name__ == "__main__":
    main()
